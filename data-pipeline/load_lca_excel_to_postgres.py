#!/usr/bin/env python3
"""Load the cleaned H-1B LCA Excel workbook into PostgreSQL.

This is the JobPush-ready loader for the canonical cleaned workbook:

    /Users/nicole/Desktop/APPLY/jobpush/LCA_H1B_FY2025_FY2026_Q2.xlsx

It keeps the existing JobLens sponsorship contract intact by rebuilding the
same FEIN-keyed tables used today:

    companies, company_aliases, company_search_keys

It also loads the full cleaned application-level fact table:

    lca_cases

and creates a conservative one-to-one default company-group mapping:

    company_groups, company_group_companies

Run with:

    DATABASE_URL=postgresql://user:pass@host:5432/db \
      python3 data-pipeline/load_lca_excel_to_postgres.py

Use --dry-run to validate and summarize the workbook without touching the DB.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from psycopg.types.json import Jsonb

import psycopg

BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent
DEFAULT_XLSX_PATH = Path("/Users/nicole/Desktop/APPLY/jobpush/LCA_H1B_FY2025_FY2026_Q2.xlsx")
SCHEMA_PATH = REPO_ROOT / "db" / "schema.sql"

DATABASE_URL = os.environ.get("DATABASE_URL")

sys.path.insert(0, str(BASE_DIR))
from export_employer_index import (  # noqa: E402
    build_index,
    compute_token_collision_counts,
    search_keys_for,
)
from naics_sectors import naics_sector_label  # noqa: E402


EXPECTED_HEADERS = [
    "CASE_NUMBER",
    "CASE_STATUS",
    "RECEIVED_DATE",
    "DECISION_DATE",
    "ORIGINAL_CERT_DATE",
    "VISA_CLASS",
    "JOB_TITLE",
    "SOC_CODE",
    "SOC_TITLE",
    "FULL_TIME_POSITION",
    "BEGIN_DATE",
    "END_DATE",
    "TOTAL_WORKER_POSITIONS",
    "NEW_EMPLOYMENT",
    "CONTINUED_EMPLOYMENT",
    "CHANGE_PREVIOUS_EMPLOYMENT",
    "NEW_CONCURRENT_EMPLOYMENT",
    "CHANGE_EMPLOYER",
    "AMENDED_PETITION",
    "EMPLOYER_NAME",
    "TRADE_NAME_DBA",
    "EMPLOYER_CITY",
    "EMPLOYER_STATE",
    "EMPLOYER_POSTAL_CODE",
    "EMPLOYER_COUNTRY",
    "EMPLOYER_PROVINCE",
    "EMPLOYER_PHONE",
    "EMPLOYER_PHONE_EXT",
    "EMPLOYER_FEIN",
    "NAICS_CODE",
    "EMPLOYER_POC_LAST_NAME",
    "EMPLOYER_POC_FIRST_NAME",
    "EMPLOYER_POC_MIDDLE_NAME",
    "EMPLOYER_POC_JOB_TITLE",
    "EMPLOYER_POC_CITY",
    "EMPLOYER_POC_STATE",
    "EMPLOYER_POC_POSTAL_CODE",
    "EMPLOYER_POC_COUNTRY",
    "EMPLOYER_POC_PROVINCE",
    "EMPLOYER_POC_PHONE",
    "EMPLOYER_POC_PHONE_EXT",
    "EMPLOYER_POC_EMAIL",
    "WORKSITE_WORKERS",
    "SECONDARY_ENTITY",
    "SECONDARY_ENTITY_BUSINESS_NAME",
    "WORKSITE_CITY",
    "WORKSITE_COUNTY",
    "WORKSITE_STATE",
    "WORKSITE_POSTAL_CODE",
    "WAGE_RATE_OF_PAY_FROM",
    "WAGE_RATE_OF_PAY_TO",
    "WAGE_UNIT_OF_PAY",
    "PREVAILING_WAGE",
    "PW_UNIT_OF_PAY",
    "PW_TRACKING_NUMBER",
    "PW_WAGE_LEVEL",
    "PW_OES_YEAR",
    "PW_OTHER_SOURCE",
    "PW_OTHER_YEAR",
    "PW_SURVEY_PUBLISHER",
    "PW_SURVEY_NAME",
    "TOTAL_WORKSITE_LOCATIONS",
    "AGREE_TO_LC_STATEMENT",
    "H_1B_DEPENDENT",
    "WILLFUL_VIOLATOR",
    "SUPPORT_H1B",
    "STATUTORY_BASIS",
    "APPENDIX_A_ATTACHED",
    "PUBLIC_DISCLOSURE",
    "PREPARER_LAST_NAME",
    "PREPARER_FIRST_NAME",
    "PREPARER_MIDDLE_INITIAL",
    "PREPARER_BUSINESS_NAME",
    "PREPARER_EMAIL",
]

HEADER_TO_COLUMN = {
    "CASE_NUMBER": "case_number",
    "CASE_STATUS": "case_status",
    "RECEIVED_DATE": "received_date",
    "DECISION_DATE": "decision_date",
    "ORIGINAL_CERT_DATE": "original_cert_date",
    "VISA_CLASS": "visa_class",
    "JOB_TITLE": "job_title",
    "SOC_CODE": "soc_code",
    "SOC_TITLE": "soc_title",
    "FULL_TIME_POSITION": "full_time_position",
    "BEGIN_DATE": "begin_date",
    "END_DATE": "end_date",
    "TOTAL_WORKER_POSITIONS": "total_worker_positions",
    "NEW_EMPLOYMENT": "new_employment",
    "CONTINUED_EMPLOYMENT": "continued_employment",
    "CHANGE_PREVIOUS_EMPLOYMENT": "change_previous_employment",
    "NEW_CONCURRENT_EMPLOYMENT": "new_concurrent_employment",
    "CHANGE_EMPLOYER": "change_employer",
    "AMENDED_PETITION": "amended_petition",
    "EMPLOYER_NAME": "employer_name",
    "TRADE_NAME_DBA": "trade_name_dba",
    "EMPLOYER_CITY": "employer_city",
    "EMPLOYER_STATE": "employer_state",
    "EMPLOYER_POSTAL_CODE": "employer_postal_code",
    "EMPLOYER_COUNTRY": "employer_country",
    "EMPLOYER_PROVINCE": "employer_province",
    "EMPLOYER_PHONE": "employer_phone",
    "EMPLOYER_PHONE_EXT": "employer_phone_ext",
    "EMPLOYER_FEIN": "employer_fein",
    "NAICS_CODE": "naics_code",
    "EMPLOYER_POC_LAST_NAME": "employer_poc_last_name",
    "EMPLOYER_POC_FIRST_NAME": "employer_poc_first_name",
    "EMPLOYER_POC_MIDDLE_NAME": "employer_poc_middle_name",
    "EMPLOYER_POC_JOB_TITLE": "employer_poc_job_title",
    "EMPLOYER_POC_CITY": "employer_poc_city",
    "EMPLOYER_POC_STATE": "employer_poc_state",
    "EMPLOYER_POC_POSTAL_CODE": "employer_poc_postal_code",
    "EMPLOYER_POC_COUNTRY": "employer_poc_country",
    "EMPLOYER_POC_PROVINCE": "employer_poc_province",
    "EMPLOYER_POC_PHONE": "employer_poc_phone",
    "EMPLOYER_POC_PHONE_EXT": "employer_poc_phone_ext",
    "EMPLOYER_POC_EMAIL": "employer_poc_email",
    "WORKSITE_WORKERS": "worksite_workers",
    "SECONDARY_ENTITY": "secondary_entity",
    "SECONDARY_ENTITY_BUSINESS_NAME": "secondary_entity_business_name",
    "WORKSITE_CITY": "worksite_city",
    "WORKSITE_COUNTY": "worksite_county",
    "WORKSITE_STATE": "worksite_state",
    "WORKSITE_POSTAL_CODE": "worksite_postal_code",
    "WAGE_RATE_OF_PAY_FROM": "wage_rate_of_pay_from",
    "WAGE_RATE_OF_PAY_TO": "wage_rate_of_pay_to",
    "WAGE_UNIT_OF_PAY": "wage_unit_of_pay",
    "PREVAILING_WAGE": "prevailing_wage",
    "PW_UNIT_OF_PAY": "pw_unit_of_pay",
    "PW_TRACKING_NUMBER": "pw_tracking_number",
    "PW_WAGE_LEVEL": "pw_wage_level",
    "PW_OES_YEAR": "pw_oes_year",
    "PW_OTHER_SOURCE": "pw_other_source",
    "PW_OTHER_YEAR": "pw_other_year",
    "PW_SURVEY_PUBLISHER": "pw_survey_publisher",
    "PW_SURVEY_NAME": "pw_survey_name",
    "TOTAL_WORKSITE_LOCATIONS": "total_worksite_locations",
    "AGREE_TO_LC_STATEMENT": "agree_to_lc_statement",
    "H_1B_DEPENDENT": "h_1b_dependent",
    "WILLFUL_VIOLATOR": "willful_violator",
    "SUPPORT_H1B": "support_h1b",
    "STATUTORY_BASIS": "statutory_basis",
    "APPENDIX_A_ATTACHED": "appendix_a_attached",
    "PUBLIC_DISCLOSURE": "public_disclosure",
    "PREPARER_LAST_NAME": "preparer_last_name",
    "PREPARER_FIRST_NAME": "preparer_first_name",
    "PREPARER_MIDDLE_INITIAL": "preparer_middle_initial",
    "PREPARER_BUSINESS_NAME": "preparer_business_name",
    "PREPARER_EMAIL": "preparer_email",
}

DATE_COLUMNS = {"received_date", "decision_date", "original_cert_date", "begin_date", "end_date"}
INTEGER_COLUMNS = {
    "total_worker_positions",
    "new_employment",
    "continued_employment",
    "change_previous_employment",
    "new_concurrent_employment",
    "change_employer",
    "amended_petition",
    "worksite_workers",
    "total_worksite_locations",
}
NUMERIC_COLUMNS = {"wage_rate_of_pay_from", "wage_rate_of_pay_to", "prevailing_wage"}

LCA_COPY_COLUMNS = [
    "source_snapshot",
    "source_row_number",
    *[HEADER_TO_COLUMN[h] for h in EXPECTED_HEADERS],
    "row_fingerprint",
]


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def normalize_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text:
        return None
    return datetime.fromisoformat(text).date()


def normalize_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(Decimal(text.replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def normalize_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    text = normalize_text(value)
    if not text:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def clean_value(column: str, value: Any) -> Any:
    if column in DATE_COLUMNS:
        return normalize_date(value)
    if column in INTEGER_COLUMNS:
        return normalize_int(value)
    if column in NUMERIC_COLUMNS:
        return normalize_decimal(value)
    return normalize_text(value)


def json_safe(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def row_fingerprint(record: dict[str, Any]) -> str:
    payload = {k: json_safe(v) for k, v in sorted(record.items())}
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def iter_clean_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if headers != EXPECTED_HEADERS:
        raise ValueError(
            "Unexpected workbook headers. Expected cleaned LCA workbook with "
            f"{len(EXPECTED_HEADERS)} columns; got {len(headers)}."
        )

    for source_row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            HEADER_TO_COLUMN[header]: clean_value(HEADER_TO_COLUMN[header], value)
            for header, value in zip(headers, values, strict=True)
        }
        yield source_row_number, record


def most_common(counter: Counter[str]) -> str | None:
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def build_employers(path: Path) -> tuple[list[dict[str, Any]], int]:
    companies: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "names": Counter(),
            "aliases": Counter(),
            "cities": Counter(),
            "states": Counter(),
            "naics": Counter(),
            "top_jobs": Counter(),
            "lca_count": 0,
            "h1b_count": 0,
            "certified_count": 0,
        }
    )
    row_count = 0

    for _, record in iter_clean_rows(path):
        row_count += 1
        fein = record["employer_fein"]
        if not fein:
            continue
        item = companies[fein]
        item["lca_count"] += 1
        if (record.get("visa_class") or "").upper() == "H-1B":
            item["h1b_count"] += 1
        if (record.get("case_status") or "").lower() == "certified":
            item["certified_count"] += 1

        for key, counter_name in (
            ("employer_name", "names"),
            ("trade_name_dba", "aliases"),
            ("employer_city", "cities"),
            ("employer_state", "states"),
            ("naics_code", "naics"),
        ):
            value = record.get(key)
            if value:
                item[counter_name][value] += 1

        title = record.get("job_title")
        if title:
            item["top_jobs"][
                (
                    title,
                    record.get("pw_wage_level"),
                    str(record.get("wage_rate_of_pay_from") or ""),
                )
            ] += 1

    raw_employers: list[dict[str, Any]] = []
    for fein, item in companies.items():
        primary_name = most_common(item["names"]) or fein
        aliases = sorted({primary_name, *item["names"].keys(), *item["aliases"].keys()})
        naics = most_common(item["naics"]) or ""
        top_jobs = [
            {
                "title": title,
                "level": level,
                "wage_from": wage_from or None,
                "count": count,
            }
            for (title, level, wage_from), count in item["top_jobs"].most_common(3)
        ]
        raw_employers.append(
            {
                "fein": fein,
                "name": primary_name,
                "names": aliases,
                "lca_count": item["lca_count"],
                "h1b_count": item["h1b_count"],
                "certified_count": item["certified_count"],
                "city": most_common(item["cities"]),
                "state": most_common(item["states"]),
                "naics_code": naics,
                "naics_sector": naics_sector_label(naics),
                "top_jobs": top_jobs,
            }
        )

    token_counts = compute_token_collision_counts([e["names"] for e in raw_employers])
    employers = [
        {
            **emp,
            "search_keys": search_keys_for(emp["name"], emp["names"], token_counts),
        }
        for emp in raw_employers
    ]
    employers.sort(key=lambda e: (-e["lca_count"], e["name"]))
    return employers, row_count


def rebuild_default_company_groups(cur) -> None:
    cur.execute(
        """
        INSERT INTO company_groups (group_key, canonical_name, display_name)
        SELECT 'legal:' || fein, name, name
        FROM companies
        ON CONFLICT (group_key) DO UPDATE
        SET canonical_name = EXCLUDED.canonical_name,
            display_name = COALESCE(company_groups.display_name, EXCLUDED.display_name),
            updated_at = now()
        """
    )
    cur.execute(
        """
        INSERT INTO company_group_companies (company_group_id, fein, relationship, is_primary)
        SELECT g.company_group_id, c.fein, 'legal_entity', TRUE
        FROM companies c
        JOIN company_groups g ON g.group_key = 'legal:' || c.fein
        ON CONFLICT (company_group_id, fein) DO UPDATE
        SET company_group_id = EXCLUDED.company_group_id,
            relationship = EXCLUDED.relationship,
            is_primary = EXCLUDED.is_primary,
            updated_at = now()
        """
    )


def upsert_companies(cur, employers: list[dict[str, Any]]) -> None:
    cur.execute(
        """
        CREATE TEMP TABLE tmp_companies (
            fein TEXT,
            name TEXT,
            naics_code TEXT,
            naics_sector TEXT,
            city TEXT,
            state TEXT,
            lca_count INTEGER,
            h1b_count INTEGER,
            certified_count INTEGER,
            top_jobs JSONB
        ) ON COMMIT DROP
        """
    )
    with cur.copy(
        "COPY tmp_companies (fein, name, naics_code, naics_sector, city, "
        "state, lca_count, h1b_count, certified_count, top_jobs) FROM STDIN"
    ) as copy:
        for e in employers:
            copy.write_row(
                [
                    e["fein"],
                    e.get("name") or "",
                    e.get("naics_code"),
                    e.get("naics_sector"),
                    e.get("city"),
                    e.get("state"),
                    e.get("lca_count") or 0,
                    e.get("h1b_count") or 0,
                    e.get("certified_count") or 0,
                    Jsonb(e.get("top_jobs") or []),
                ]
            )
    cur.execute(
        """
        UPDATE companies
        SET lca_count = 0,
            h1b_count = 0,
            certified_count = 0,
            top_jobs = '[]'::jsonb
        """
    )
    cur.execute(
        """
        INSERT INTO companies (
            fein, name, naics_code, naics_sector, city, state,
            lca_count, h1b_count, certified_count, top_jobs
        )
        SELECT
            fein, name, naics_code, naics_sector, city, state,
            lca_count, h1b_count, certified_count, top_jobs
        FROM tmp_companies
        ON CONFLICT (fein) DO UPDATE
        SET name = EXCLUDED.name,
            naics_code = EXCLUDED.naics_code,
            naics_sector = EXCLUDED.naics_sector,
            city = EXCLUDED.city,
            state = EXCLUDED.state,
            lca_count = EXCLUDED.lca_count,
            h1b_count = EXCLUDED.h1b_count,
            certified_count = EXCLUDED.certified_count,
            top_jobs = EXCLUDED.top_jobs
        """
    )


def copy_aliases(cur, employers: list[dict[str, Any]]) -> None:
    with cur.copy("COPY company_aliases (fein, alias_name) FROM STDIN") as copy:
        for e in employers:
            seen: set[str] = set()
            for alias in [e.get("name"), *(e.get("names") or [])]:
                if not alias or alias in seen:
                    continue
                seen.add(alias)
                copy.write_row([e["fein"], alias])


def copy_search_keys(cur, employers: list[dict[str, Any]]) -> None:
    key_index = build_index(employers)
    with cur.copy("COPY company_search_keys (search_key, fein) FROM STDIN") as copy:
        for key, employer in key_index.items():
            copy.write_row([key, employer["fein"]])


def copy_lca_cases(cur, path: Path, source_snapshot: str) -> int:
    sql = f"COPY lca_cases ({', '.join(LCA_COPY_COLUMNS)}) FROM STDIN"
    count = 0
    with cur.copy(sql) as copy:
        for source_row_number, record in iter_clean_rows(path):
            count += 1
            fingerprint = row_fingerprint(record)
            copy.write_row(
                [
                    source_snapshot,
                    source_row_number,
                    *[record[HEADER_TO_COLUMN[h]] for h in EXPECTED_HEADERS],
                    fingerprint,
                ]
            )
            if count % 100_000 == 0:
                print(f"  copied {count:,} LCA rows")
    return count


def load_to_postgres(path: Path, source_snapshot: str) -> None:
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is required unless --dry-run is used.")

    t0 = time.time()
    print(f"Scanning workbook: {path}")
    employers, row_count = build_employers(path)
    print(f"  {row_count:,} LCA rows")
    print(f"  {len(employers):,} FEIN-keyed companies")

    with psycopg.connect(DATABASE_URL) as conn:
        with conn.cursor() as cur:
            print("Applying schema")
            cur.execute(SCHEMA_PATH.read_text(encoding="utf-8"))

            print("Clearing replaceable derived/base tables")
            cur.execute("TRUNCATE lca_cases RESTART IDENTITY")
            cur.execute("TRUNCATE company_aliases RESTART IDENTITY")
            cur.execute("TRUNCATE company_search_keys")

            print("Upserting companies")
            upsert_companies(cur, employers)

            print("Loading aliases")
            copy_aliases(cur, employers)

            print("Loading search keys")
            copy_search_keys(cur, employers)

            print("Rebuilding default company groups")
            rebuild_default_company_groups(cur)

            print("Loading full LCA cases")
            copied = copy_lca_cases(cur, path, source_snapshot)
            if copied != row_count:
                raise RuntimeError(f"Copied {copied:,} rows, expected {row_count:,}")

        conn.commit()

        with conn.cursor() as cur:
            counts = {}
            for table in (
                "lca_cases",
                "companies",
                "company_aliases",
                "company_search_keys",
                "company_groups",
                "company_group_companies",
            ):
                cur.execute(f"SELECT count(*) FROM {table}")
                counts[table] = cur.fetchone()[0]

    print(f"Done in {time.time() - t0:.1f}s")
    for table, n in counts.items():
        print(f"  {table}: {n:,} rows")


def dry_run(path: Path) -> None:
    t0 = time.time()
    employers, row_count = build_employers(path)
    key_index = build_index(employers)
    print(f"Workbook: {path}")
    print(f"Rows: {row_count:,}")
    print(f"Companies by FEIN: {len(employers):,}")
    print(f"Search keys: {len(key_index):,}")
    print("Top 10 companies:")
    for i, emp in enumerate(employers[:10], start=1):
        print(f"  {i:>2}. {emp['name']} ({emp['fein']}): {emp['lca_count']:,}")
    print(f"Dry run completed in {time.time() - t0:.1f}s")


def main() -> None:
    parser = argparse.ArgumentParser(description="Load cleaned H-1B LCA Excel into Postgres")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX_PATH)
    parser.add_argument("--source-snapshot", default="LCA_H1B_FY2025_FY2026_Q2.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(args.xlsx)

    if args.dry_run:
        dry_run(args.xlsx)
    else:
        load_to_postgres(args.xlsx, args.source_snapshot)


if __name__ == "__main__":
    main()
